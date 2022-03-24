import openpyxl
path=r"E:\ExcelFiles\file_example_XLSX_10.xlsx"
wb=openpyxl.load_workbook(path)
sheet=wb.active
rows=sheet.max_row
cols=sheet.max_column
print(rows,cols)
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end='    ')
    print("")